from pathlib import Path
from typing import Generator

import pandas as pd
from openpyxl import load_workbook, Workbook

from constants import APP_CONFIG_FILE, APP_CONFIG_MAIN_SETTINGS_SHEET, APP_CONFIG_URLS_TO_PROTOCOLS_SHEET, \
    CONFIG_VERSION_SHEET, APP_CONFIG_VERSION
from errors import Error, AppConfigValidationError


def check_app_config():
    check_app_config_exists()
    wb: Workbook = load_workbook(APP_CONFIG_FILE, read_only=True, keep_vba=False)
    check_workbook(wb)


def check_app_config_exists():
    if not Path(APP_CONFIG_FILE).exists():
        raise Error('Excel file with application configuration was not found.')


def check_workbook(wb: Workbook):
    if CONFIG_VERSION_SHEET not in wb.sheetnames:
        raise AppConfigValidationError('Sheet with version was not found.')
    else:
        version, = next(wb[CONFIG_VERSION_SHEET].values)
        if version != APP_CONFIG_VERSION:
            raise AppConfigValidationError(
                f'This version is not supported by the tool. Supported version is {APP_CONFIG_VERSION} but you are '
                f'using {version}.')

    if APP_CONFIG_MAIN_SETTINGS_SHEET not in wb.sheetnames:
        raise AppConfigValidationError(
            f'Sheet "{APP_CONFIG_MAIN_SETTINGS_SHEET}" was not found.')
    main_settings = dict(value[:2] for value in wb[APP_CONFIG_MAIN_SETTINGS_SHEET].values)

    if 'Тип источника протоколов' not in main_settings:
        raise AppConfigValidationError('Field "Тип источника протоколов" was not found.')
    protocol_source_type = main_settings['Тип источника протоколов']
    if protocol_source_type not in ('Файл', 'Ссылка'):
        raise AppConfigValidationError(f'Unsupported protocol source type: "{protocol_source_type}".')

    for rank_name in (
    'Общего летнего ранга', 'Общего зимнего ранга', 'Лесного ранга', 'Спринт ранга', 'гонки сильнейших'):
        protocols_dir = f'Путь к папке со всеми протоколами для {rank_name}'
        if protocols_dir not in main_settings:
            raise AppConfigValidationError(f'Field "{protocols_dir}" was not found.')
        protocols_dir = Path(main_settings[protocols_dir])
        if protocol_source_type == 'Файл':
            if not protocols_dir.is_dir():
                raise AppConfigValidationError(f'Directory with protocols does not exist: "{protocols_dir}".')
            # if not any(protocols_dir.glob('*.htm')):
            #     raise AppConfigValidationError(
            #         f'Directory "{protocols_dir}" does not contain any protocols (.htm files).')

    if protocol_source_type == 'Ссылка':
        check_urls_to_protocols(wb)

    # TODO: Check 'Путь к файлу с финальным рангом предыдущего года'
    # TODO: Check 'Путь к папке с результатами'


def check_urls_to_protocols(wb: Workbook):
    if APP_CONFIG_URLS_TO_PROTOCOLS_SHEET not in wb.sheetnames:
        raise AppConfigValidationError(
            f'Sheet "{APP_CONFIG_URLS_TO_PROTOCOLS_SHEET}" was not found.')

    values: Generator = wb[APP_CONFIG_URLS_TO_PROTOCOLS_SHEET].values
    header = next(values)
    if header != ('№ п/п', 'Ссылка', 'Учитывать при расчете ранга для Гонки сильнейших'):
        raise AppConfigValidationError(
            f'Unsupported header on the "{APP_CONFIG_URLS_TO_PROTOCOLS_SHEET}" sheet: {header}.')

    df = pd.DataFrame(values, columns=header)
    if df.empty:
        raise AppConfigValidationError(f'The "{APP_CONFIG_URLS_TO_PROTOCOLS_SHEET}" sheet does not contain any rows.')

    for column in ('№ п/п', 'Ссылка'):
        if not df[column].is_unique:
            raise AppConfigValidationError(f'The column "{column}" contains duplicated values on the '
                                           f'"{APP_CONFIG_URLS_TO_PROTOCOLS_SHEET}" sheet.')
