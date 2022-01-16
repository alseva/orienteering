from pathlib import Path

from openpyxl import load_workbook, Workbook

from constants import RANK_CONFIG_FILE, RANK_CONFIG_MAIN_SETTINGS_SHEET, RANK_CONFIG_RACE_TYPE_SHEET, \
    RANK_CONFIG_RACE_LEVEL_SHEET, RANK_CONFIG_GROUP_RANK_SHEET, RANK_CONFIG_PENALTY_LACK_RACES_SHEET, \
    RANK_CONFIG_PENALTY_NOT_STARTED_SHEET, RANK_CONFIG_PENALTY_LEFT_RACE_SHEET, CONFIG_VERSION_SHEET, \
    RANK_CONFIG_VERSION
from errors import Error, RankConfigValidationError


def check_rank_config():
    check_rank_config_exists()
    wb: Workbook = load_workbook(RANK_CONFIG_FILE, read_only=True, keep_vba=False)
    check_workbook(wb)


def check_rank_config_exists():
    if not Path(RANK_CONFIG_FILE).exists():
        raise Error('Excel file with rank configuration was not found.')


def check_workbook(wb: Workbook):
    if CONFIG_VERSION_SHEET not in wb.sheetnames:
        raise RankConfigValidationError('Sheet with version was not found.')
    else:
        version, = next(wb[CONFIG_VERSION_SHEET].values)
        if version != RANK_CONFIG_VERSION:
            raise RankConfigValidationError(
                f'This version is not supported by the tool. Supported version is {RANK_CONFIG_VERSION} but you are '
                f'using {version}.')

    for sheet in (RANK_CONFIG_MAIN_SETTINGS_SHEET,
                  RANK_CONFIG_RACE_TYPE_SHEET,
                  RANK_CONFIG_RACE_LEVEL_SHEET,
                  RANK_CONFIG_GROUP_RANK_SHEET,
                  RANK_CONFIG_PENALTY_LACK_RACES_SHEET,
                  RANK_CONFIG_PENALTY_NOT_STARTED_SHEET,
                  RANK_CONFIG_PENALTY_LEFT_RACE_SHEET):
        if sheet not in wb.sheetnames:
            raise RankConfigValidationError(f'Sheet "{sheet}" was not found.')
